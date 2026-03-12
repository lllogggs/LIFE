from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


def _parse_csv(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)

    if not rows:
        raise ValueError("csv file is empty")

    return {
        "raw_fields": {key: str(value) for key, value in rows[0].items() if key},
        "rows": rows,
        "confidence": 0.8,
        "source_meta": {"source_type": "csv", "file_path": str(path), "row_count": len(rows)},
    }


def _parse_xlsx_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("xlsx parsing requires openpyxl. install openpyxl or use csv") from exc

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("xlsx file is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        row_dict: dict[str, Any] = {}
        for idx, value in enumerate(row):
            key = header[idx] if idx < len(header) and header[idx] else f"col_{idx+1}"
            row_dict[key] = value
        if any(v is not None and str(v).strip() != "" for v in row_dict.values()):
            data_rows.append(row_dict)

    if not data_rows:
        raise ValueError("xlsx file has no data rows")

    return {
        "raw_fields": {k: str(v) for k, v in data_rows[0].items()},
        "rows": data_rows,
        "confidence": 0.8,
        "source_meta": {"source_type": "xlsx", "file_path": str(path), "row_count": len(data_rows)},
    }


def parse_excel_like(file_path: str) -> dict[str, Any]:
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"spreadsheet file not found: {file_path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        return _parse_csv(path)
    if ext in {".xlsx", ".xlsm", ".xltx"}:
        return _parse_xlsx_with_openpyxl(path)

    raise ValueError(f"unsupported spreadsheet extension: {ext}")
